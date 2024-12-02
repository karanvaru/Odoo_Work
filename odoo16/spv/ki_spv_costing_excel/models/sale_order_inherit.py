# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
import base64
import io
from openpyxl import Workbook, load_workbook
from odoo.exceptions import UserError
from xlsx2html import xlsx2html
import pdfkit
import subprocess
from openpyxl.cell.cell import MergedCell


class SaleOrder(models.Model):
    _inherit = 'sale.order'

    distributor_id = fields.Many2one(
        'res.partner',
        string="Distributor"
    )
    panel_distributor_id = fields.Many2one(
        'res.partner',
        string="Panel Distributor"
    )
    area_type = fields.Selection([
        ('urban', 'Urban'),
        ('rural', 'Rural'),
    ],
        string='Area Type'
    )
    terrif_type = fields.Selection([
        ('residential', 'Residential'),
        ('commercial', 'Commercial'),
    ],
        string='Terrif Type'
    )
    discom = fields.Selection([
        ('TPL', 'TPL'),
        ('UGVCL', 'UGVCL'),
        ('PGVCL', 'PGVCL'),
        ('DGVCL', 'DGVCL'),
        ('MGVCL', 'MGVCL'),
    ],
        string='Discom'
    )
    kw_capacity = fields.Float(
        'KW Capacity'
    )
    # internal_note = fields.Text(
    #     "Internal Note"
    # )

    # Excel Sheet

    # original_file = fields.Binary(
    #     string='Original File',
    # )
    # original_char = fields.Char(
    #     string='Original File'
    # )
    generated_file = fields.Binary(
        string='Generated File',
    )
    generated_char = fields.Char(
        string='Generated File'
    )
    generated_pdf_file = fields.Binary(
        string='Generated PDF File',
    )
    generated_pdf_char = fields.Char(
        string='Generated PDF File'
    )

    def copy_sheet_with_styles(self, source_sheet, target_sheet):
        for row in source_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()

                if source_sheet.row_dimensions[cell.row].height:
                    target_sheet.row_dimensions[cell.row].height = source_sheet.row_dimensions[cell.row].height
                if not isinstance(cell, MergedCell) and cell.column_letter in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[cell.column_letter].width = source_sheet.column_dimensions[
                        cell.column_letter].width

        if source_sheet.merged_cells:
            for merged_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_range))

    def generate_file(self):
        template = self.env['quote.excel.template'].search([], limit=1)

        if not template.file:
            raise UserError(_('Set Sample File Before Generate File!'))

        file_data = base64.b64decode(template.file)
        file_like = io.BytesIO(file_data)
        wb = load_workbook(file_like)

        self._update_excel_data(wb, template)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        modified_file_content = base64.b64encode(output.read())
        self.generated_file = modified_file_content
        self.generated_char = 'Generated_Report.xlsx'

        generated_data = base64.b64decode(self.generated_file)
        generated_file_like = io.BytesIO(generated_data)
        generated_wb = load_workbook(generated_file_like)

        new_wb = Workbook()
        new_wb.remove(new_wb.active)

        first_sheet = generated_wb.worksheets[0]
        new_sheet = new_wb.create_sheet(title=first_sheet.title)

        self.copy_sheet_with_styles(first_sheet, new_sheet)

        new_output = io.BytesIO()
        new_wb.save(new_output)
        new_output.seek(0)

        self._generate_pdf_from_excel(new_output)

    def _update_excel_data(self, wb, template):
        for rec in template.line_ids:
            sheets = wb.worksheets
            if rec.sheet_number > len(sheets):
                raise UserError(_('Added Sheet Number Is Not Available In Current File'))
            sheet = sheets[rec.sheet_number - 1]

            if rec.field_id.model_id.name == 'Sales Order':
                field_value = self.sudo().read([rec.field_id.name], load=False)
                if rec.field_id.name == 'terrif_type':
                    field_value[0].update({
                        'terrif_type': 'Tarrif Type: {}'.format(self.terrif_type)
                    })
                if rec.field_id.name == 'discom':
                    field_value[0].update({
                        'discom': 'Discom: {}'.format(self.discom)
                    })

            if rec.field_id.model_id.name == 'Contact':
                field_value = self.partner_id.sudo().read([rec.field_id.name], load=False)
                if rec.field_id.name == 'display_name':
                    field_value[0].update({
                        'display_name': 'To:- {}'.format(self.partner_id.display_name)
                    })
                if rec.field_id.name == 'city':
                    field_value[0].update({
                        'city': 'Location:- {}'.format(self.partner_id.city)
                    })

            if rec.field_id.model_id.name == 'Sales Order Line':
                if not self.order_line:
                    raise UserError(_('At Least One Order Line Required!'))
                field_value = self.order_line[0].sudo().read([rec.field_id.name], load=False)
            # print("___________    field_value", field_value)
            if sheet.merged_cells:
                for merged_cell in sheet.merged_cells.ranges:
                    if rec.cell in merged_cell:
                        sheet.cell(merged_cell.min_row, merged_cell.min_col).value = list(field_value[-1].values())[-1]
                        break
            else:
                sheet[rec.cell] = list(field_value[-1].values())[-1]














        # first_sheet['H8'] = self.date_order
        # first_sheet['B9'] = 'To:- {}'.format(self.partner_id.display_name)
        # first_sheet['B10'] = 'Location:- {}'.format(self.partner_id.city)
        # terrif_type_value = 'Tarrif Type: {}'.format(self.terrif_type)
        # discom_value = 'Discom: {}'.format(self.discom)
        # if first_sheet.merged_cells:
        #     for merged_cell in first_sheet.merged_cells.ranges:
        #         if 'F10' in merged_cell:
        #             first_sheet.cell(merged_cell.min_row, merged_cell.min_col).value = terrif_type_value
        #             break
        # else:
        #     first_sheet['F10'] = terrif_type_value

        # if first_sheet.merged_cells:
        #     for merged_cell in first_sheet.merged_cells.ranges:
        #         if 'H10' in merged_cell:
        #             first_sheet.cell(merged_cell.min_row, merged_cell.min_col).value = discom_value
        #             break
        # else:
        #     first_sheet['H10'] = discom_value

        # if self.order_line:
        #     first_sheet['B12'] = self.order_line[0].name
        #     first_sheet['F12'] = self.order_line[0].product_uom_qty
        #     first_sheet['G12'] = self.order_line[0].price_unit
        #     first_sheet['H12'] = self.order_line[0].price_subtotal

    def adjust_column_widths(self, excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active

        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 2

        wb.save(excel_file_path)

    def _generate_pdf_from_excel(self, output):
        output.seek(0)
        file_data = output.read()
        excel_file_path = '/tmp/Generated_Report.xlsx'
        with open(excel_file_path, 'wb') as f:
            f.write(file_data)

        self.adjust_column_widths(excel_file_path)

        html_file_path = '/tmp/Generated_Report.html'
        xlsx2html(excel_file_path, html_file_path)

        pdf_file_path = '/tmp/Generated_Report.pdf'
        pdfkit.from_file(html_file_path, pdf_file_path)

        subprocess.run(['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', '/tmp', excel_file_path])

        with open(pdf_file_path, 'rb') as f:
            pdf_data = f.read()
            self.generated_pdf_file = base64.b64encode(pdf_data)
            self.generated_pdf_char = 'Generated_Report.pdf'

        return True
