from odoo import models, fields, api, _
import base64
import io
from openpyxl import Workbook, load_workbook
import zipfile
import xlrd
from odoo.exceptions import UserError
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


class EstimateExcel(models.Model):
    _name = 'estimate.excel'

    original_file = fields.Binary(
        string='Original File',
    )
    original_char = fields.Char(
        string='Original File'
    )
    generated_file = fields.Binary(
        string='Generated File',
    )
    generated_char = fields.Char(
        string='Generated File'
    )
    generated_pdf_file = fields.Binary(
        string='Generated PDF File',
    )
    generated_pdf_char = fields.Char(
        string='Generated PDF File'
    )
    estimate_excel_line_ids = fields.One2many(
        "estimate.excel.line",
        "estimate_excel_id",
        string="Lines"
    )

    def generate_file(self):
        if not self.original_file:
            raise UserError(_('Select Original File Before Generate File!'))

        file_data = base64.b64decode(self.original_file)
        file_like = io.BytesIO(file_data)
        try:
            wb = load_workbook(file_like)

        # not .xlsx files
        except zipfile.BadZipFile:
            file_like.seek(0)
            book = xlrd.open_workbook(file_contents=file_data)
            sheet = book.sheet_by_index(0)

            wb = Workbook()
            ws = wb.active

            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    ws.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

        cell_values = {}
        ws = wb.active
        for rec in self.estimate_excel_line_ids:
            ws[rec.cell] = rec.data
            cell_reference_data = ws[rec.cell].coordinate
            column_letter = ''.join(filter(str.isalpha, cell_reference_data))
            if column_letter not in cell_values:
                cell_values[column_letter] = {}
            cell_values[column_letter].update({
                cell_reference_data: rec.data
            })
        for rec in cell_values:
            positive_data = 0
            negative_data = 0
            last_key = list(cell_values[rec].keys())[-1]
            for val in cell_values[rec]:
                if cell_values[rec][val] <= 0:
                    negative_data += cell_values[rec][val]
                if cell_values[rec][val] >= 0:
                    positive_data += cell_values[rec][val]
            cell_reference = ws[last_key].coordinate
            column_letter = ''.join(filter(str.isalpha, cell_reference))
            row_number = ''.join(filter(str.isdigit, cell_reference))
            next_row_number = int(row_number) + 1
            next_cell = f'{column_letter}{next_row_number}'
            if negative_data <= 0:
                ws[next_cell] = positive_data + negative_data
            if negative_data >= 0:
                ws[next_cell] = positive_data - negative_data

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        modified_file_content = base64.b64encode(output.read())
        self.generated_file = modified_file_content
        self.generated_char = 'Generated_Report.xlsx'

        pdf_content = self._generate_pdf_from_excel(wb)
        self.generated_pdf_file = base64.b64encode(pdf_content)
        self.generated_pdf_char = 'Generated_Report.pdf'

    def _generate_pdf_from_excel(self, workbook):
        pdf_output = io.BytesIO()
        pdf = canvas.Canvas(pdf_output, pagesize=A4)
        pdf.setFont("Helvetica", 10)
        width, height = A4
        margin = 100
        x_offset = margin
        y_offset = height - margin

        ws = workbook.active
        pdf.drawString(x_offset, y_offset, "Data from Excel:")
        y_offset -= 20

        for row in ws.iter_rows(values_only=True):
            row_data = "           ".join([str(cell) if cell is not None else '           ' for cell in row])
            pdf.drawString(x_offset, y_offset, row_data)
            y_offset -= 15

            if y_offset < margin:
                pdf.showPage()
                y_offset = height - margin

        pdf.showPage()
        pdf.save()

        pdf_output.seek(0)
        return pdf_output.getvalue()
